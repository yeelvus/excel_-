#!/usr/bin/env python3
"""
Excel翻译替换工具
输入目录: INPUT_DIR 中所有 .xlsx 文件
翻译对照: TRANS_JSON (original/translation 键值对列表)
输出目录: OUTPUT_DIR

规则：
- 跳过公式单元格（值以 '=' 开头）
- 对纯文本单元格做精确匹配替换（去首尾空格后比较）
- 合并单元格的占位格（MergedCell）跳过，不修改
"""

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

INPUT_DIR = Path('/Users/wyattsun/企业微信/WXWork Files/Caches/Files/2026-05/bff19a7dbab56c802f4de931eaff67ff')
TRANS_JSON = Path('/Users/wyattsun/Projects/excel翻译工具/翻译对照.json')
OUTPUT_DIR = Path('/Users/wyattsun/Projects/excel翻译工具/4_输出文件excel')


def build_lookup(path: Path) -> dict[str, str]:
    data = json.loads(path.read_text(encoding='utf-8'))
    lookup: dict[str, str] = {}
    for item in data:
        orig = item.get('original', '').strip()
        trans = item.get('translation', '').strip()
        if orig and trans:
            lookup[orig] = trans
    return lookup


def replace_multiline_text(text: str, lookup: dict[str, str]) -> tuple[str, int]:
    normalized = text.replace('\r\n', '\n').replace('\r', '\n')
    parts = normalized.split('\n')
    replaced = 0
    new_parts: list[str] = []

    for part in parts:
        stripped = part.strip()
        if stripped and stripped in lookup:
            leading = len(part) - len(part.lstrip())
            trailing = len(part) - len(part.rstrip())
            prefix = part[:leading]
            suffix = part[len(part) - trailing:] if trailing else ''
            new_parts.append(f"{prefix}{lookup[stripped]}{suffix}")
            replaced += 1
        else:
            new_parts.append(part)

    if replaced == 0:
        return text, 0
    return '\n'.join(new_parts), replaced


def translate_workbook(src: Path, dst: Path, lookup: dict[str, str]) -> int:
    wb = load_workbook(src, data_only=False)
    count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val is None:
                    continue
                # 跳过公式单元格
                if isinstance(val, str) and val.lstrip().startswith('='):
                    continue
                # 字符串单元格做精确匹配替换
                if isinstance(val, str):
                    replaced_text, replaced_count = replace_multiline_text(val, lookup)
                    if replaced_count > 0:
                        cell.value = replaced_text
                        count += replaced_count
    wb.save(dst)
    return count


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    lookup = build_lookup(TRANS_JSON)
    print(f"已加载翻译词条: {len(lookup)} 条")

    xlsx_files = sorted(INPUT_DIR.glob('*.xlsx'))
    print(f"发现 Excel 文件: {len(xlsx_files)} 个\n")

    total = 0
    for src in xlsx_files:
        dst = OUTPUT_DIR / src.name
        n = translate_workbook(src, dst, lookup)
        total += n
        print(f"  [{n:>4} 处替换]  {src.name}")

    print(f"\n完成！共替换 {total} 处，输出目录: {OUTPUT_DIR}")


if __name__ == '__main__':
    main()

