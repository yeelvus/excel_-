#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from decimal import Decimal
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def split_cell_text_lines(text: str) -> list[str]:
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    return normalized.split("\n")


PURE_NUMBER_RE = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?$")
INDEX_NUMBER_RE = re.compile(r"^\d+(?:\.\d+)+$")
SHORT_ALNUM_CODE_RE = re.compile(r"^\d+[A-Za-z]{1,3}$")
LETTER_DIGIT_CODE_RE = re.compile(r"^[A-Za-z]{1,6}\d+(?:\.\d+)*$")


def is_pure_numeric_cell(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float, Decimal)):
        return True
    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return False
        normalized = text.replace(",", "")
        return bool(PURE_NUMBER_RE.fullmatch(normalized))
    return False


def is_index_or_code_cell(value: Any) -> bool:
    if not isinstance(value, str):
        return False

    text = value.strip()
    if text == "":
        return False

    compact = "".join(text.split())
    candidate = compact.strip(";,，；、。:：")
    if candidate == "":
        return False

    if INDEX_NUMBER_RE.fullmatch(candidate):
        return True
    if SHORT_ALNUM_CODE_RE.fullmatch(candidate):
        return True
    if LETTER_DIGIT_CODE_RE.fullmatch(candidate):
        return True
    return False


def extract_excel_to_txt(
    excel_path: Path,
    txt_path: Path,
    include_empty: bool = False,
    with_sheet_header: bool = False,
    dedupe_text: bool = True,
) -> int:
    workbook = load_workbook(filename=excel_path, data_only=True, read_only=True)
    lines_written = 0

    txt_path.parent.mkdir(parents=True, exist_ok=True)

    with txt_path.open("w", encoding="utf-8", newline="\n") as out_file:
        total_sheets = len(workbook.worksheets)
        seen_texts: set[str] = set()

        for index, ws in enumerate(workbook.worksheets, start=1):
            if with_sheet_header and total_sheets > 1:
                out_file.write(f"=== Sheet: {ws.title} ===\n")

            for row in ws.iter_rows(values_only=True):
                for value in row:
                    if value is None:
                        if include_empty:
                            out_file.write("\n")
                            lines_written += 1
                        continue

                    if is_pure_numeric_cell(value):
                        continue

                    if is_index_or_code_cell(value):
                        continue

                    text = cell_to_text(value)
                    if text == "":
                        continue

                    for line_text in split_cell_text_lines(text):
                        line_text = line_text.strip()

                        if line_text == "":
                            if include_empty:
                                out_file.write("\n")
                                lines_written += 1
                            continue

                        if dedupe_text and line_text in seen_texts:
                            continue

                        if dedupe_text:
                            seen_texts.add(line_text)

                        out_file.write(line_text + "\n")
                        lines_written += 1

            if with_sheet_header and total_sheets > 1 and index != total_sheets:
                out_file.write("\n")

    workbook.close()
    return lines_written


BASE_DIR = Path(__file__).resolve().parent
EXCEL_SUFFIXES = {".xlsx", ".xls", ".xlsm", ".xlsb"}
OUTPUT_DIR = BASE_DIR / "1_提取文本"
MERGE_DIR = BASE_DIR / "2_提取文件合并文件"
TRANSLATION_JSON_CANDIDATES = [
    BASE_DIR / "翻译对照.json",
    BASE_DIR / "4_输出文件excel" / "翻译对照.json",
]

# ── 智能过滤正则 ──
CHINESE_RE = re.compile(r"[\u4e00-\u9fff]")
THAI_RE = re.compile(r"[\u0e00-\u0e7f]")
# 含有3个以上连续英文字母的单词（说明有实际英文内容）
ENGLISH_WORD_RE = re.compile(r"[A-Za-z]{3,}")

# 需要剔除的行模式（纯数字/纯符号/纯代码，无有意义的文字）
JUNK_LINE_RE = re.compile(
    r"^[\s\-]*("
    # 纯数字（含千分位逗号、小数、负号、百分号）
    r"[+-]?[\d,]+(?:\.\d+)?%?"
    r"|"
    # 纯序号格式: 1. / 1.1 / 1.1.1
    r"(?:\d+\.)+\d*"
    r"|"
    # 纯设备编号/面板代码（如 A-R01-1DP1, 93A-1-CBP-01, B-R02-MDB1）
    r"[A-Z0-9][-A-Z0-9_./,()]*(?:\s+to\s+[A-Z0-9][-A-Z0-9_./,()]*)?\.?"
    r"|"
    # 纯尺寸/规格（如 Dia. 1 1/2", -Size 300 sq.mm., 400 A）
    r"(?:Dia\.?|Size)\s*[\d\s/\"x.]+(?:sq\.?\s*mm\.?|mm\.?)?"
    r"|"
    # 纯数值+单位（如 400 A, 20 A, 100FU, 160FU）
    r"[\d,]+(?:\.\d+)?\s*(?:A|FU|kVA|kW|hp|V|W|mm\.?|m\.?|sq\.?\s*mm\.?)\.?"
    r"|"
    # 纯符号 / 纯标点
    r'[-=_~*#/\\|:;,.!?@$%^&()\[\]{}<>"' "'" r"`，。；：、！？（）【】《》\u201c\u201d\u2018\u2019…—\s]+"
    r"|"
    # 短代码（如 93A, 93B, LS, INC, SUM, ADD, N/A, #REF!）
    r"[A-Z0-9]{1,5}[./]?[A-Z0-9]{0,3}"
    r"|"
    # Excel错误值（如 #REF!, #N/A, #VALUE!, #DIV/0!）
    r"#[A-Z/0-9]+!?"
    r")[\s.]*$",
    re.IGNORECASE,
)


def collect_excel_files(input_path: Path) -> list[Path]:
    if input_path.is_file():
        if input_path.suffix.lower() in EXCEL_SUFFIXES:
            return [input_path]
        else:
            print(f"跳过非Excel文件: {input_path.name}")
            return []
    elif input_path.is_dir():
        files = sorted(
            f for f in input_path.rglob("*")
            if f.suffix.lower() in EXCEL_SUFFIXES and not f.name.startswith("~$")
        )
        return files
    else:
        raise FileNotFoundError(f"路径不存在: {input_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="提取Excel单元格文本内容到md文件（支持文件或文件夹）"
    )
    parser.add_argument(
        "input_path",
        nargs="?",
        help="Excel文件路径或包含Excel文件的文件夹路径",
    )
    parser.add_argument(
        "--include-empty",
        action="store_true",
        help="输出空单元格为空行",
    )
    parser.add_argument(
        "--with-sheet-header",
        action="store_true",
        help="在输出中添加Sheet标题分隔",
    )
    parser.add_argument(
        "--no-dedupe",
        action="store_true",
        help="保留重复文本行（默认去重）",
    )
    return parser.parse_args()


def normalize_path_text(path_text: str) -> str:
    return path_text.strip().strip('"').strip("'")


def main() -> None:
    args = parse_args()

    raw_input = args.input_path
    if raw_input is None:
        raw_input = input("请输入Excel文件或文件夹路径: ").strip()

    raw_input = normalize_path_text(raw_input)
    if not raw_input:
        raise ValueError("路径不能为空")

    input_path = Path(raw_input).expanduser().resolve()
    excel_files = collect_excel_files(input_path)

    if not excel_files:
        print("未找到任何Excel文件")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    total_files = len(excel_files)
    total_lines = 0

    print(f"找到 {total_files} 个Excel文件，输出目录: {OUTPUT_DIR}\n")

    for i, excel_path in enumerate(excel_files, start=1):
        md_name = excel_path.stem + ".md"
        md_path = OUTPUT_DIR / md_name

        # 避免同名覆盖：加上父目录名
        if md_path.exists():
            md_name = f"{excel_path.parent.name}_{excel_path.stem}.md"
            md_path = OUTPUT_DIR / md_name

        try:
            lines = extract_excel_to_txt(
                excel_path=excel_path,
                txt_path=md_path,
                include_empty=args.include_empty,
                with_sheet_header=args.with_sheet_header,
                dedupe_text=not args.no_dedupe,
            )
            total_lines += lines
            print(f"[{i}/{total_files}] {excel_path.name} → {md_name} ({lines} 行)")
        except Exception as e:
            print(f"[{i}/{total_files}] {excel_path.name} 失败: {e}")

    print(f"\n完成！共处理 {total_files} 个文件，提取 {total_lines} 行文本到 {OUTPUT_DIR}")

    # 合并所有md文件并去重
    merge_md_files()


def merge_md_files() -> None:
    md_files = sorted(OUTPUT_DIR.glob("*.md"))
    if not md_files:
        print("\n没有可合并的md文件")
        return

    MERGE_DIR.mkdir(parents=True, exist_ok=True)
    merged_path = MERGE_DIR / "合并文本.md"

    seen: set[str] = set()
    unique_lines: list[str] = []
    total_raw = 0

    for md_file in md_files:
        for line in md_file.read_text(encoding="utf-8").splitlines():
            total_raw += 1
            stripped = line.strip()
            if stripped == "":
                continue
            if stripped in seen:
                continue
            seen.add(stripped)
            unique_lines.append(line)

    merged_path.write_text("\n".join(unique_lines) + "\n", encoding="utf-8")
    print(f"\n合并完成: {len(md_files)} 个文件, 原始 {total_raw} 行 → 去重后 {len(unique_lines)} 行")
    print(f"输出: {merged_path}")

    # 智能过滤
    smart_filter(merged_path)


def normalize_match_text(text: str) -> str:
    text = text.strip()
    text = re.sub(r"\s+", " ", text)
    return text


def find_translation_json_path() -> Path:
    for candidate in TRANSLATION_JSON_CANDIDATES:
        if candidate.exists():
            return candidate
    return TRANSLATION_JSON_CANDIDATES[0]


def load_translation_originals(json_path: Path) -> tuple[set[str], set[str]]:
    if not json_path.exists():
        return set(), set()

    data = json.loads(json_path.read_text(encoding="utf-8"))
    originals: list[str] = []

    if isinstance(data, list):
        for item in data:
            if not isinstance(item, dict):
                continue
            original = item.get("original")
            if isinstance(original, str) and original.strip():
                originals.append(original.strip())
    elif isinstance(data, dict):
        for key in data:
            if isinstance(key, str) and key.strip():
                originals.append(key.strip())

    exact_set = set(originals)
    normalized_set = {normalize_match_text(text) for text in originals}
    return exact_set, normalized_set


def export_pending_translation_items(filtered_lines: list[str]) -> None:
    pending_path = MERGE_DIR / "待翻译项.md"

    if not filtered_lines:
        pending_path.write_text("", encoding="utf-8")
        print("待翻译项: 0 行")
        print(f"输出: {pending_path}")
        return

    translation_json_path = find_translation_json_path()
    exact_originals, normalized_originals = load_translation_originals(translation_json_path)
    if not exact_originals and not normalized_originals:
        pending_lines = filtered_lines
        print(f"未找到翻译对照文件或文件为空: {translation_json_path}")
    else:
        pending_lines = []
        for line in filtered_lines:
            normalized_line = normalize_match_text(line)
            if line in exact_originals or normalized_line in normalized_originals:
                continue
            pending_lines.append(line)

    pending_path.write_text("\n".join(pending_lines) + ("\n" if pending_lines else ""), encoding="utf-8")
    print(f"待翻译项: {len(pending_lines)} 行")
    print(f"输出: {pending_path}")


def _chinese_ratio(text: str) -> float:
    """计算中文字符占非空白字符的比例。"""
    non_space = re.sub(r"\s", "", text)
    if not non_space:
        return 0.0
    cn_count = len(CHINESE_RE.findall(non_space))
    return cn_count / len(non_space)


def should_keep_line(line: str) -> tuple[bool, str]:
    """判断一行是否应保留，返回 (保留?, 剔除原因)。
    保留：含泰文、含有意义的英文描述
    剔除：纯中文（含中文为主+少量英文缩写）、纯数字/符号/设备编号
    """
    stripped = line.strip().lstrip("- ")
    if not stripped:
        return False, "空行"

    has_chinese = bool(CHINESE_RE.search(stripped))
    has_thai = bool(THAI_RE.search(stripped))
    has_english_word = bool(ENGLISH_WORD_RE.search(stripped))

    # 含泰文 → 保留
    if has_thai:
        return True, ""

    # 含中文的行：看中文占比，超过30%视为中文主体 → 剔除
    if has_chinese and _chinese_ratio(stripped) > 0.3:
        return False, "纯中文"

    # 纯中文（无英文）→ 剔除
    if has_chinese and not has_english_word:
        return False, "纯中文"

    # 以下是无中文无泰文的行，判断是否有意义的英文
    # 有英文单词且不是纯代码/纯数字 → 保留
    if has_english_word and not JUNK_LINE_RE.fullmatch(stripped):
        return True, ""

    # 纯数字/代码/符号 → 剔除
    return False, "数字/代码/符号"


def smart_filter(merged_path: Path) -> None:
    """从合并文件中提取含泰文的有效行，纯中文和其他剔除行分别存档。"""
    if not merged_path.exists():
        return

    lines = merged_path.read_text(encoding="utf-8").splitlines()
    kept: list[str] = []
    removed_chinese: list[str] = []
    removed_other: list[str] = []

    for ln in lines:
        keep, reason = should_keep_line(ln)
        if keep:
            kept.append(ln)
        elif reason == "纯中文":
            removed_chinese.append(ln)
        elif reason != "空行":
            removed_other.append(ln)

    filtered_path = MERGE_DIR / "过滤后文本.md"
    filtered_path.write_text("\n".join(kept) + "\n", encoding="utf-8")

    chinese_path = MERGE_DIR / "剔除_纯中文.md"
    chinese_path.write_text("\n".join(removed_chinese) + "\n", encoding="utf-8")

    other_path = MERGE_DIR / "剔除_数字单位代码等.md"
    other_path.write_text("\n".join(removed_other) + "\n", encoding="utf-8")

    total = len(lines)
    print(f"\n智能过滤: {total} 行 → 保留 {len(kept)} 行")
    print(f"  剔除纯中文: {len(removed_chinese)} 行 → {chinese_path.name}")
    print(f"  剔除数字/单位/代码/符号/纯英文: {len(removed_other)} 行 → {other_path.name}")
    print(f"输出: {filtered_path}")

    export_pending_translation_items(kept)


if __name__ == "__main__":
    main()
