#!/usr/bin/env python3
from __future__ import annotations

import argparse
from bisect import bisect_left
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
UNIT_HEADER_RE = re.compile(r"^(单位|單位|unit|uom|qty\s*unit|หน่วย)$", re.IGNORECASE)
DEFAULT_INPUT_DIR = Path("/Users/wyattsun/Desktop/CPE清单翻译")
DEFAULT_OUTPUT_DIR = Path("/Users/wyattsun/Projects/excel翻译工具/4_输出文件excel-去重复")


@dataclass
class SheetStats:
    sheet_name: str
    merged_lines: int = 0
    removed_rows: int = 0
    unit_col: int = 0


def normalize_path_text(path_text: str) -> str:
    return path_text.strip().strip('"').strip("'")


def is_target_sheet(sheet_name: str) -> bool:
    upper = sheet_name.upper()
    return "SUM" not in upper


def is_empty_cell(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def to_clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def collect_excel_files(input_path: Path) -> list[Path]:
    if input_path.is_file():
        if input_path.suffix.lower() in EXCEL_SUFFIXES and not input_path.name.startswith("~$"):
            return [input_path]
        return []

    if input_path.is_dir():
        return sorted(
            f
            for f in input_path.rglob("*")
            if f.suffix.lower() in EXCEL_SUFFIXES and not f.name.startswith("~$")
        )

    raise FileNotFoundError(f"路径不存在: {input_path}")


def find_unit_col_by_header(ws, header_row: int, max_scan_col: int = 80) -> int | None:
    for col in range(1, max_scan_col + 1):
        val = ws.cell(row=header_row, column=col).value
        text = to_clean_text(val)
        if text and UNIT_HEADER_RE.search(text):
            return col
    return None


def append_detail_text(base_text: str, extra_text: str) -> str:
    if not base_text:
        return extra_text
    if not extra_text:
        return base_text
    return f"{base_text}\n{extra_text}"


def _capture_sheet_formulas(ws) -> list[tuple[int, int, str]]:
    formulas: list[tuple[int, int, str]] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas.append((cell.row, cell.column, cell.value))
    return formulas


def _restore_shifted_formulas(ws, formulas: list[tuple[int, int, str]], deleted_rows: list[int]) -> None:
    if not formulas or not deleted_rows:
        return

    deleted_rows_sorted = sorted(deleted_rows)
    deleted_set = set(deleted_rows_sorted)

    for old_row, col, old_formula in formulas:
        if old_row in deleted_set:
            continue

        # 计算删除后新行号：减去 old_row 之前被删除的行数。
        shift = bisect_left(deleted_rows_sorted, old_row)
        new_row = old_row - shift

        old_coord = f"{get_column_letter(col)}{old_row}"
        new_coord = f"{get_column_letter(col)}{new_row}"

        try:
            new_formula = Translator(old_formula, origin=old_coord).translate_formula(new_coord)
        except Exception:
            new_formula = old_formula

        ws.cell(row=new_row, column=col).value = new_formula


def process_sheet(
    ws,
    header_row: int,
    data_start_row: int,
    seq_col: int,
    second_col: int,
    detail_col: int,
    unit_col: int,
    delete_merged_rows: bool,
) -> SheetStats:
    stats = SheetStats(sheet_name=ws.title, unit_col=unit_col)

    last_main_row: int | None = None
    rows_to_delete: list[int] = []

    formulas_before_delete = _capture_sheet_formulas(ws) if delete_merged_rows else []

    for r in range(data_start_row, ws.max_row + 1):
        seq_val = ws.cell(r, seq_col).value
        b_val = ws.cell(r, second_col).value
        unit_val = ws.cell(r, unit_col).value
        detail_val = ws.cell(r, detail_col).value

        detail_text = to_clean_text(detail_val)
        is_continuation = (
            is_empty_cell(seq_val)
            and is_empty_cell(b_val)
            and is_empty_cell(unit_val)
            and detail_text != ""
        )

        if is_continuation and last_main_row is not None:
            main_detail_cell = ws.cell(last_main_row, detail_col)
            main_text = to_clean_text(main_detail_cell.value)
            main_detail_cell.value = append_detail_text(main_text, detail_text)
            stats.merged_lines += 1
            if delete_merged_rows:
                rows_to_delete.append(r)
            continue

        row_has_any_content = False
        for c in range(1, ws.max_column + 1):
            if not is_empty_cell(ws.cell(r, c).value):
                row_has_any_content = True
                break

        if row_has_any_content:
            last_main_row = r

    if delete_merged_rows and rows_to_delete:
        for r in reversed(rows_to_delete):
            ws.delete_rows(r, 1)
        stats.removed_rows = len(rows_to_delete)
        _restore_shifted_formulas(ws, formulas_before_delete, rows_to_delete)

    return stats


def process_workbook(
    src_path: Path,
    dst_path: Path,
    header_row: int,
    data_start_row: int,
    seq_col: int,
    second_col: int,
    detail_col: int,
    unit_col: int | None,
    delete_merged_rows: bool,
) -> tuple[int, int, list[SheetStats]]:
    wb = load_workbook(src_path)
    target_sheets = [s for s in wb.sheetnames if is_target_sheet(s)]

    stats_list: list[SheetStats] = []
    total_merged = 0
    total_removed = 0

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        effective_unit_col = unit_col
        if effective_unit_col is None:
            effective_unit_col = find_unit_col_by_header(ws, header_row) or 6

        stats = process_sheet(
            ws=ws,
            header_row=header_row,
            data_start_row=data_start_row,
            seq_col=seq_col,
            second_col=second_col,
            detail_col=detail_col,
            unit_col=effective_unit_col,
            delete_merged_rows=delete_merged_rows,
        )
        stats_list.append(stats)
        total_merged += stats.merged_lines
        total_removed += stats.removed_rows

    dst_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst_path)
    wb.close()
    return total_merged, total_removed, stats_list


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="BOQ跨行描述合并：识别三空续行并合并到上一条主记录"
    )
    parser.add_argument(
        "input_path",
        nargs="?",
        default=str(DEFAULT_INPUT_DIR),
        help="Excel文件或目录，默认使用脚本内置输入目录",
    )
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="输出目录，默认使用脚本内置输出目录",
    )
    parser.add_argument("--header-row", type=int, default=1, help="表头行号，默认 1")
    parser.add_argument(
        "--data-start-row",
        type=int,
        default=2,
        help="数据起始行号，默认 2",
    )
    parser.add_argument("--seq-col", type=int, default=1, help="序号列，默认 A=1")
    parser.add_argument("--second-col", type=int, default=2, help="第二列，默认 B=2")
    parser.add_argument("--detail-col", type=int, default=3, help="明细列，默认 C=3")
    parser.add_argument(
        "--unit-col",
        type=int,
        default=None,
        help="单位列。默认自动识别表头(单位/unit/uom/หน่วย)，识别不到回退 F=6",
    )
    parser.add_argument(
        "--keep-continuation-rows",
        action="store_true",
        help="仅合并文本，不删除续行",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    input_path = Path(normalize_path_text(args.input_path)).expanduser().resolve()
    output_dir = Path(normalize_path_text(args.output_dir)).expanduser().resolve()

    excel_files = collect_excel_files(input_path)
    if not excel_files:
        print("未找到可处理的Excel文件")
        return

    print(f"找到 {len(excel_files)} 个Excel文件")
    print("处理所有工作表，仅排除名称包含 SUM 的工作表\n")

    grand_merged = 0
    grand_removed = 0

    for i, src in enumerate(excel_files, start=1):
        dst = output_dir / src.name
        merged, removed, stats_list = process_workbook(
            src_path=src,
            dst_path=dst,
            header_row=args.header_row,
            data_start_row=args.data_start_row,
            seq_col=args.seq_col,
            second_col=args.second_col,
            detail_col=args.detail_col,
            unit_col=args.unit_col,
            delete_merged_rows=not args.keep_continuation_rows,
        )
        grand_merged += merged
        grand_removed += removed

        print(f"[{i}/{len(excel_files)}] {src.name}")
        if not stats_list:
            print("  无匹配Sheet，已原样输出")
        else:
            for st in stats_list:
                print(
                    f"  - {st.sheet_name}: 合并 {st.merged_lines} 行, "
                    f"删除 {st.removed_rows} 行, 单位列={st.unit_col}"
                )
        print(f"  输出: {dst}\n")

    print("处理完成")
    print(f"总合并行数: {grand_merged}")
    print(f"总删除续行: {grand_removed}")
    print(f"输出目录: {output_dir}")


if __name__ == "__main__":
    main()
