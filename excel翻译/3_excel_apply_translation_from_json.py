#!/usr/bin/env python3
"""
Excel翻译替换工具
输入目录: INPUT_DIR 中所有 .xlsx 文件
翻译对照: TRANS_JSON (original/translation 键值对列表)
输出目录: OUTPUT_DIR

规则：
- 跳过公式单元格（值以 '=' 开头）
- 对纯文本单元格做精确匹配替换（去首尾空格后比较）
- 合并单元格的占位格（MergedCell）跳过，不修改
"""

import json
import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

BASE_DIR = Path(__file__).resolve().parent
EXCEL_SUFFIXES = {'.xlsx', '.xlsm', '.xltx', '.xltm'}
TRANS_JSON = BASE_DIR / '翻译对照.json'
OUTPUT_DIR = BASE_DIR / '4_输出文件excel'


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description='根据翻译对照json批量输出已翻译Excel'
    )
    parser.add_argument(
        'input_path',
        nargs='?',
        help='待输出翻译的Excel文件或文件夹路径',
    )
    parser.add_argument(
        '--json-path',
        default=str(TRANS_JSON),
        help='翻译对照json路径',
    )
    parser.add_argument(
        '--output-dir',
        default=str(OUTPUT_DIR),
        help='输出目录',
    )
    return parser.parse_args()


def normalize_path_text(path_text: str) -> str:
    return path_text.strip().strip('"').strip("'")


def collect_excel_files(input_path: Path) -> list[Path]:
    if input_path.is_file():
        if input_path.suffix.lower() in EXCEL_SUFFIXES and not input_path.name.startswith('~$'):
            return [input_path]
        print(f'跳过非Excel文件: {input_path.name}')
        return []
    if input_path.is_dir():
        return sorted(
            file
            for file in input_path.rglob('*')
            if file.suffix.lower() in EXCEL_SUFFIXES and not file.name.startswith('~$')
        )
    raise FileNotFoundError(f'路径不存在: {input_path}')


def build_lookup(path: Path) -> dict[str, str]:
    data = json.loads(path.read_text(encoding='utf-8'))
    lookup: dict[str, str] = {}
    if isinstance(data, dict):
        for orig, trans in data.items():
            if isinstance(orig, str) and isinstance(trans, str):
                if orig.strip() and trans.strip():
                    lookup[orig.strip()] = trans.strip()
        return lookup

    for item in data:
        orig = item.get('original', '').strip()
        trans = item.get('translation', '').strip()
        if orig and trans:
            lookup[orig] = trans
    return lookup


def replace_multiline_text(text: str, lookup: dict[str, str]) -> tuple[str, int]:
    normalized = text.replace('\r\n', '\n').replace('\r', '\n')
    parts = normalized.split('\n')
    replaced = 0
    new_parts: list[str] = []

    for part in parts:
        stripped = part.strip()
        if stripped and stripped in lookup:
            leading = len(part) - len(part.lstrip())
            trailing = len(part) - len(part.rstrip())
            prefix = part[:leading]
            suffix = part[len(part) - trailing:] if trailing else ''
            new_parts.append(f"{prefix}{lookup[stripped]}{suffix}")
            replaced += 1
        else:
            new_parts.append(part)

    if replaced == 0:
        return text, 0
    return '\n'.join(new_parts), replaced


def translate_workbook(src: Path, dst: Path, lookup: dict[str, str]) -> int:
    wb = load_workbook(src, data_only=False)
    count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val is None:
                    continue
                # 跳过公式单元格
                if isinstance(val, str) and val.lstrip().startswith('='):
                    continue
                # 字符串单元格做精确匹配替换
                if isinstance(val, str):
                    replaced_text, replaced_count = replace_multiline_text(val, lookup)
                    if replaced_count > 0:
                        cell.value = replaced_text
                        count += replaced_count
    wb.save(dst)
    return count


def main() -> None:
    args = parse_args()

    raw_input = args.input_path
    if raw_input is None:
        raw_input = input('请输入Excel文件或文件夹路径: ').strip()

    raw_input = normalize_path_text(raw_input)
    if not raw_input:
        raise ValueError('路径不能为空')

    input_path = Path(raw_input).expanduser().resolve()
    json_path = Path(args.json_path).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()

    excel_files = collect_excel_files(input_path)
    if not excel_files:
        print('未找到任何Excel文件')
        return

    output_dir.mkdir(parents=True, exist_ok=True)
    lookup = build_lookup(json_path)
    print(f"已加载翻译词条: {len(lookup)} 条")

    source_root = input_path if input_path.is_dir() else input_path.parent
    print(f"发现 Excel 文件: {len(excel_files)} 个\n")

    total = 0
    for src in excel_files:
        relative_path = src.relative_to(source_root) if input_path.is_dir() else Path(src.name)
        dst = output_dir / relative_path
        dst.parent.mkdir(parents=True, exist_ok=True)
        n = translate_workbook(src, dst, lookup)
        total += n
        print(f"  [{n:>4} 处替换]  {relative_path}")

    print(f"\n完成！共替换 {total} 处，输出目录: {output_dir}")


if __name__ == '__main__':
    main()

